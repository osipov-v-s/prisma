"""XLSX exports with one observation per row and no merged research cells."""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def _finish(workbook: Workbook) -> bytes:
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="294A8A")
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(50, max(12, *(len(str(cell.value or "")) + 2 for cell in column)))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _append_responses(sheet, session: dict) -> None:
    sheet.append(["session_id", "user_name", "collection", "is_training", "presentation_index",
                  "level_index", "left_item_id", "left_type_id", "left_type_name",
                  "right_item_id", "right_type_id", "right_type_name",
                  "selected_item_id", "selected_type_id", "selected_type_name",
                  "reaction_time_ms", "exceeded_time_limit",
                  "timed_out", "status", "shown_at", "answered_at"])
    for item in session.get("comparisons") or []:
        sheet.append([session["id"], session["user_name"], session["collection_name"], item["is_training"],
                      item["presentation_index"], item["level_index"], item["left_item_id"], item["left_type_id"],
                      item["left_type_name"], item["right_item_id"], item["right_type_id"], item["right_type_name"],
                      item["selected_item_id"], item["selected_type_id"], item["selected_type_name"],
                      item["reaction_time_ms"], item["exceeded_time_limit"], item["timed_out"], item["status"],
                      item["shown_at"], item["answered_at"]])


def build_session_xlsx(session: dict) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["field", "value"])
    for key in ("id", "user_name", "collection_name", "started_at", "finished_at", "status",
                "time_mode", "time_limit_ms", "random_seed"):
        summary.append([key, session.get(key)])
    analysis = session.get("analysis") or {}
    for mode in ("choice_only", "choice_and_time"):
        for score in analysis.get(mode, {}).get("overall") or []:
            summary.append([f'{mode}:{score["type_name"]}', score["percent"]])
    _append_responses(workbook.create_sheet("Responses"), session)
    levels = workbook.create_sheet("LevelResults")
    levels.append(["analysis_mode", "pair_first_type", "pair_second_type", "level_index",
                   "observation_role", "selected_type_id", "reaction_time_ms"])
    for mode in ("choice_only", "choice_and_time"):
        for pair in analysis.get(mode, {}).get("pair_aggregations") or []:
            for observation in pair.get("observations") or []:
                levels.append([mode, pair["first_type_id"], pair["second_type_id"], observation["level_index"],
                               observation["role"], observation["selected_type_id"], observation["reaction_time_ms"]])
    metadata = workbook.create_sheet("AlgorithmMetadata")
    metadata.append(["field", "value"])
    metadata.append(["algorithm_version", analysis.get("algorithm_version")])
    for mode in ("choice_only", "choice_and_time"):
        item = analysis.get(mode, {})
        for key in ("status", "time_algorithm", "iteration_strategy", "coverage", "epsilon", "max_iterations"):
            metadata.append([f"{mode}.{key}", str(item.get(key))])
    return _finish(workbook)


def build_admin_xlsx(sessions: list[dict]) -> bytes:
    workbook = Workbook()
    users = workbook.active
    users.title = "Users"
    users.append(["account_id", "full_name"])
    seen = set()
    for session in sessions:
        if session["account_id"] not in seen:
            users.append([session["account_id"], session["user_name"]])
            seen.add(session["account_id"])
    session_sheet = workbook.create_sheet("Sessions")
    session_sheet.append(["session_id", "account_id", "user_name", "collection_id", "collection_name",
                          "started_at", "finished_at", "status", "time_mode", "time_limit_ms", "random_seed"])
    response_sheet = workbook.create_sheet("Responses")
    first = True
    score_sheet = workbook.create_sheet("Scores")
    score_sheet.append(["session_id", "analysis_mode", "type_id", "type_name", "score", "percent"])
    level_sheet = workbook.create_sheet("LevelScores")
    level_sheet.append(["session_id", "analysis_mode", "pair_first_type", "pair_second_type",
                        "level_index", "role", "selected_type_id", "reaction_time_ms"])
    metadata = workbook.create_sheet("AlgorithmMetadata")
    metadata.append(["session_id", "algorithm_version", "mode", "status", "algorithm", "coverage", "epsilon"])
    for session in sessions:
        session_sheet.append([session[key] for key in ("id", "account_id", "user_name", "collection_id",
                              "collection_name", "started_at", "finished_at", "status", "time_mode",
                              "time_limit_ms", "random_seed")])
        temporary = Workbook(); sheet = temporary.active
        _append_responses(sheet, session)
        if first:
            response_sheet.append([cell.value for cell in sheet[1]])
            first = False
        for row in sheet.iter_rows(min_row=2, values_only=True): response_sheet.append(row)
        analysis = session.get("analysis") or {}
        for mode in ("choice_only", "choice_and_time"):
            value = analysis.get(mode, {})
            for score in value.get("overall") or []:
                score_sheet.append([session["id"], mode, score["type_id"], score["type_name"], score["score"], score["percent"]])
            for pair in value.get("pair_aggregations") or []:
                for observation in pair.get("observations") or []:
                    level_sheet.append([session["id"], mode, pair["first_type_id"], pair["second_type_id"],
                                        observation["level_index"], observation["role"],
                                        observation["selected_type_id"], observation["reaction_time_ms"]])
            metadata.append([session["id"], analysis.get("algorithm_version"), mode, value.get("status"),
                             value.get("time_algorithm"), value.get("coverage"), value.get("epsilon")])
    return _finish(workbook)
