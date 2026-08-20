"""Complete collection passage, analytics, reports, and reproducible schedule."""

import base64
from io import BytesIO

from openpyxl import load_workbook

from src.service import auth, experiments, reports


def test_collection_starts_without_mandatory_training(desktop_db) -> None:
    account = auth.get_account(next(item["id"] for item in auth.list_users() if item["login"] == "user"))
    test = experiments.available_tests()[0]
    session = experiments.create_session(account.id, test["id"], "fixed-seed")

    assert session["status"] == "in_progress"
    assert session["training_total"] == 0
    assert session["main_total"] == 30

    initial_trace = experiments.get_session(session["id"], account.id, False, True)["comparisons"]
    assert [item["level_index"] for item in initial_trace] == [
        level for level in range(1, 6) for _ in range(6)
    ]

    seen_pairs = set()
    type_order: dict[str, int] = {}
    collection = next(item for item in experiments.list_collections() if item["id"] == test["id"])
    item_levels = {
        cell["item_id"]: cell["level_index"]
        for row in collection["rows"]
        for cell in row["cells"]
    }
    while session["status"] == "in_progress":
        session = experiments.present_next(session["id"], account.id)
        pair = session["next_comparison"]
        type_order.setdefault(pair["left_type_id"], len(type_order))
        type_order.setdefault(pair["right_type_id"], len(type_order))
        key = (pair["level_index"], *sorted((pair["left_type_id"], pair["right_type_id"])))
        assert key not in seen_pairs
        assert item_levels[pair["left_item_id"]] == pair["level_index"]
        assert item_levels[pair["right_item_id"]] == pair["level_index"]
        seen_pairs.add(key)
        selected = (pair["left_item_id"] if type_order[pair["left_type_id"]] <
                    type_order[pair["right_type_id"]] else pair["right_item_id"])
        session = experiments.save_response(
            session["id"], pair["presentation_index"], account.id,
            {"selected_item_id": selected, "reaction_time_ms": 100 + pair["presentation_index"]},
        )

    assert session["status"] == "completed"
    assert session["analysis"]["algorithm_version"].startswith("prisma-analytics/")
    assert len(seen_pairs) == 30
    assert base64.b64decode(reports.session_pdf(session["id"], account)).startswith(b"%PDF")
    xlsx = base64.b64decode(reports.session_xlsx(session["id"], account))
    assert xlsx.startswith(b"PK")
    workbook = load_workbook(BytesIO(xlsx), read_only=True, data_only=True)
    assert workbook.sheetnames == [
        "Summary", "Responses", "Matrices", "LevelResults", "AlgorithmMetadata"
    ]
    assert workbook["Matrices"].max_row > 1


def test_same_seed_produces_same_pair_order(desktop_db) -> None:
    account = auth.get_account(next(item["id"] for item in auth.list_users() if item["login"] == "user"))
    collection_id = experiments.available_tests()[0]["id"]
    first = experiments.create_session(account.id, collection_id, "same")
    second = experiments.create_session(account.id, collection_id, "same")
    first_trace = experiments.get_session(first["id"], account.id, False, True)["comparisons"]
    second_trace = experiments.get_session(second["id"], account.id, False, True)["comparisons"]
    projection = lambda trace: [(item["left_item_id"], item["right_item_id"]) for item in trace]
    assert projection(first_trace) == projection(second_trace)
